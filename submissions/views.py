import logging
from collections import OrderedDict
from typing import Any, Dict, Tuple
from decimal import Decimal, InvalidOperation
import re
import copy
import json
import hashlib

import openpyxl
from django.contrib import messages
from django import forms
from django.forms import formset_factory
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from users.models import User
from .parsers import (
    NTPDParserUnavailable,
    get_ntpd_sheet_specs,
    parse_with_ntpd,
)
from .forms import (
    SubmissionForm,
    MultiUploadForm,
    SubmissionEventForm,
    SubmissionDynamicFieldsForm,
)
from .models import Submission, Upload, SubmissionEvent, SubmissionField

logger = logging.getLogger(__name__)


def _is_lender_like(user: User) -> bool:
    """
    대주 권한으로 인정할 기준:
    - role == LENDER 이거나
    - is_staff / is_superuser 인 관리자
    """
    return getattr(user, "is_lender", False) or user.is_staff or user.is_superuser


# -------------------------------
# 파일 파싱 → SubmissionField 생성 관련 헬퍼
# -------------------------------

def _auto_name_from_label(label: str) -> str:
    """
    라벨에서 자동으로 name을 만들어내는 간단한 함수.
    - 한글/영문/숫자 외 문자는 제거 또는 언더스코어 대체
    - 공백은 언더스코어로
    """
    s = str(label).strip()
    # 공백을 언더스코어로
    s = re.sub(r"\s+", "_", s)
    # 너무 지저분하면 그냥 field_... 형태로 가도 됨
    return s


def extract_fields_from_upload(upload: Upload):
    """
    업로드된 엑셀 파일을 읽어서
    SubmissionField로 쓸 원시 데이터 리스트를 반환한다.

    현재 가정하는 엑셀 구조 (예시):

        A열: 라벨 (예: "당기순이익", "자산총계")
        B열: 값   (예: 999, 88888)

    이 함수는 엑셀에 몇 행이 있든,
    라벨/값이 차있는 만큼 모두 자동으로 필드를 생성한다.

    반환 형식:

        [
            {
                "name": "당기순이익",   # 라벨 기반 자동 생성
                "label": "당기순이익",
                "value": "999",
                "data_type": SubmissionField.DataType.NUMBER,
                "order": 1,
            },
            ...
        ]
    """
    file_path = upload.file.path  # 실제 서버 상의 파일 경로

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    fields = []
    order = 1

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        label_cell, value_cell = row

        if label_cell is None or value_cell is None:
            continue

        label_str = str(label_cell).strip()
        value_str = str(value_cell).strip()

        # name 자동 생성 (label 기반)
        name = _auto_name_from_label(label_str)

        # 숫자인지 간단 판별
        data_type = SubmissionField.DataType.TEXT
        try:
            Decimal(value_str.replace(",", ""))
            data_type = SubmissionField.DataType.NUMBER
        except InvalidOperation:
            pass

        fields.append(
            {
                "name": name,
                "label": label_str,
                "value": value_str,
                "data_type": data_type,
                "order": order,
            }
        )
        order += 1

    return fields


def _records_to_field_specs(records: list[dict]) -> list[dict]:
    """
    ntpd extractor records -> SubmissionField 생성용 dict 리스트
    """
    specs = []
    for idx, record in enumerate(records, start=1):
        key_tokens = [token for token in record.get("key_path", []) if token]
        label_tokens = [token for token in record.get("label_path", []) if token]
        raw_value = record.get("value")
        value_str = "" if raw_value is None else str(raw_value)

        data_type = SubmissionField.DataType.TEXT
        try:
            Decimal(value_str.replace(",", ""))
            data_type = SubmissionField.DataType.NUMBER
        except (InvalidOperation, AttributeError):
            pass

        specs.append(
            {
                "name": ".".join(key_tokens) if key_tokens else f"field_{idx}",
                "label": " / ".join(label_tokens) if label_tokens else f"Field {idx}",
                "value": value_str,
                "data_type": data_type,
                "order": idx,
            }
        )
    return specs


def _rebuild_submission_fields(submission: Submission, records: list[dict]) -> None:
    specs = _records_to_field_specs(records)
    SubmissionField.objects.filter(submission=submission).delete()
    objs = [
        SubmissionField(
            submission=submission,
            name=spec.get("name") or f"field_{idx}",
            label=spec.get("label") or f"Field {idx}",
            value=spec.get("value") or "",
            data_type=spec.get("data_type") or SubmissionField.DataType.TEXT,
            order=spec.get("order") or idx,
        )
        for idx, spec in enumerate(specs, start=1)
    ]
    if objs:
        SubmissionField.objects.bulk_create(objs)


def _format_display_value(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value if v not in (None, ""))
    return value


def _as_list_items(value: Any):
    if value is None:
        return None
    if isinstance(value, (list, tuple)):
        items = [str(v).strip() for v in value if v not in (None, "")]
        return items or None
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    prepared = text.replace(", -", "\n-").replace(",  -", "\n-")
    lines = []
    for line in prepared.splitlines():
        part = line.strip()
        if not part:
            continue
        if part.startswith("-"):
            part = part[1:].strip()
        lines.append(part)
    return lines or None


def _is_numeric(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    try:
        text = str(value).replace(",", "").strip()
        if text == "":
            return False
        Decimal(text)
        return True
    except Exception:
        return False


def _resolve_path(data: Any, path: list[str] | None):
    if not path:
        return None
    cur = data
    for key in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(key)
        if cur is None:
            return None
    return cur


def _key_path_to_str(path: list[str] | None) -> str:
    if not path:
        return ""
    return "|".join(str(p) for p in path)


def _apply_overrides_to_records(records: list[dict], overrides: dict[str, Any]) -> list[dict]:
    if not overrides:
        return records

    updated: list[dict] = []
    for record in records:
        key = _key_path_to_str(record.get("key_path"))
        new_value = overrides.get(key, record.get("value"))
        updated.append(
            {
                "key_path": record.get("key_path") or [],
                "label_path": record.get("label_path") or [],
                "value": new_value,
            }
        )
    return updated


def _clone_payload(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        return copy.deepcopy(payload)
    except Exception:
        return json.loads(json.dumps(payload))


def _apply_overrides_to_payload(payload: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    if not overrides:
        return payload

    cloned = _clone_payload(payload)
    for key_str, value in overrides.items():
        path = [segment for segment in key_str.split("|") if segment]
        if not path:
            continue
        _set_nested_value(cloned, path, value)
    return cloned


def _set_nested_value(data: dict[str, Any], path: list[str], value: Any) -> None:
    cur = data
    for key in path[:-1]:
        if not isinstance(cur, dict):
            return
        cur = cur.get(key)
        if cur is None:
            return
    if isinstance(cur, dict):
        cur[path[-1]] = value


def _get_effective_records(submission: Submission) -> list[dict]:
    records = submission.normalized_records or []
    overrides = submission.normalized_overrides or {}
    if overrides and records:
        return _apply_overrides_to_records(records, overrides)
    return records


def _get_effective_payload(submission: Submission) -> dict[str, Any]:
    payload = submission.normalized_payload or {}
    overrides = submission.normalized_overrides or {}
    if overrides and payload:
        return _apply_overrides_to_payload(payload, overrides)
    return payload


def _build_spec_views_from_payload(payload: dict[str, Any]):
    sheet_specs = get_ntpd_sheet_specs()
    views: list[dict] = []

    for sheet_name, specs in sheet_specs.items():
        sheet_payload = payload.get(sheet_name)
        if not sheet_payload:
            continue

        sections: list[dict] = []
        basic_rows: list[dict] = []

        for spec in specs:
            mode = spec.get("mode")
            if mode == "table":
                block = _build_plain_table_section(sheet_payload, spec)
                if block:
                    sections.append(block)
                continue

            if mode == "period_table":
                blocks = _build_period_table_sections(sheet_payload, spec)
                if blocks:
                    sections.extend(blocks)
                continue

            value = _resolve_path(sheet_payload, spec.get("key_path"))
            if value in (None, "", []):
                continue
            label = spec.get("label_path", [sheet_name, "항목"])[-1]
            basic_rows.append(
                {
                    "label": label,
                    "value": _format_display_value(value),
                    "items": _as_list_items(value),
                    "align": "right" if _is_numeric(value) else "left",
                    "raw_value": str(value or ""),
                    "key": _key_path_to_str(spec.get("key_path")),
                }
            )

        if basic_rows:
            sections.insert(
                0,
                {
                    "mode": "list",
                    "title": "기본 항목",
                    "rows": basic_rows,
                },
            )

        if sections:
            views.append({"sheet": sheet_name, "sections": sections})

    return views


def _build_plain_table_section(sheet_payload: dict[str, Any], spec: dict[str, Any]):
    rows_data = _resolve_path(sheet_payload, spec.get("parent_key_path"))
    if not rows_data:
        return None

    columns_map = spec.get("columns_map") or {}
    col_keys = [cfg[0] for cfg in columns_map.values()]
    headers = [cfg[1] for cfg in columns_map.values()]

    table_rows = []
    for row in rows_data:
        row_cells = []
        for key in col_keys:
            raw = (row or {}).get(key, "")
            row_cells.append(
                {
                    "value": _format_display_value(raw),
                }
            )
        table_rows.append(row_cells)

    total = None
    total_label = None
    total_path = spec.get("total_key_path")
    if total_path:
        total_row = _resolve_path(sheet_payload, total_path)
        if isinstance(total_row, dict):
            total = []
            for key in col_keys:
                raw = total_row.get(key, "")
                total.append(
                    {
                        "value": _format_display_value(raw),
                        "align": "right" if _is_numeric(raw) else "left",
                    }
                )
            total_label_path = spec.get("total_label_path")
            if total_label_path:
                total_label = total_label_path[-1]
                if total:
                    total = total[1:]

    title = spec.get("parent_label_path", ["표 데이터"])[-1]
    return {
        "mode": "plain_table",
        "title": title,
        "headers": headers,
        "rows": table_rows,
        "total": total,
        "total_label": total_label,
    }


def _build_period_table_sections(sheet_payload: dict[str, Any], spec: dict[str, Any]):
    table_root = _resolve_path(sheet_payload, spec.get("sheet_key_path"))
    if not table_root:
        return []

    row_defs = _period_row_definitions(spec, table_root)
    columns, has_subheaders = _period_columns(table_root, spec, row_defs)
    if not columns:
        return []

    rows = []
    for row_def in row_defs:
        entry = _period_entry_data(table_root, row_def.get("data_path"))
        cells = _build_period_row_cells(entry, columns)
        rows.append(
            {
                "label": row_def["label"],
                "indent": row_def.get("indent", 0),
                "is_group": row_def.get("is_group", False),
                "indent_px": row_def.get("indent", 0) * 18,
                "cells": cells,
            }
        )

    title = spec.get("sheet_label_path", ["기간 테이블"])[-1]
    return [
        {
            "mode": "period_table",
            "title": title,
            "columns": columns,
            "has_subheaders": has_subheaders,
            "rows": rows,
        }
    ]


def _period_row_definitions(spec: dict[str, Any], table_root: dict[str, Any]):
    defs: list[dict] = []

    row_tree = spec.get("row_tree")
    if row_tree:
        for label, node in row_tree.items():
            defs.extend(_expand_row_tree(label, node, 0))
        return defs

    row_key_map = spec.get("row_key_map")
    if row_key_map:
        for label, key in row_key_map.items():
            defs.append({"label": label, "data_path": (key,)})
        return defs

    for key in table_root.keys():
        defs.append({"label": key, "data_path": (key,)})
    return defs


def _expand_row_tree(label: str, node: Any, indent: int):
    rows = []
    if isinstance(node, dict) and "children" in node:
        rows.append({"label": label, "data_path": None, "indent": indent, "is_group": True})
        parent_key = node.get("key")
        children = node.get("children") or {}
        for child_label, child_key in children.items():
            rows.append(
                {
                    "label": child_label,
                    "data_path": (parent_key, child_key),
                    "indent": indent + 1,
                }
            )
    else:
        key = node.get("key") if isinstance(node, dict) else node
        rows.append({"label": label, "data_path": (key,), "indent": indent})
    return rows


def _period_columns(table_root: dict[str, Any], spec: dict[str, Any], row_defs: list[dict]):
    period_order: list[str] = []
    col_depth = spec.get("col_header_depth", 1)

    for row_def in row_defs:
        entry = _period_entry_data(table_root, row_def.get("data_path"))
        if not isinstance(entry, dict):
            continue
        for period in entry.keys():
            if period not in period_order:
                period_order.append(period)

    if not period_order:
        return [], False

    columns = []
    has_subheaders = False
    col_key_map = spec.get("col_key_map")

    if col_depth >= 2 and col_key_map:
        has_subheaders = True
        subheaders = [{"label": label, "key": key} for label, key in col_key_map.items()]
        for period in period_order:
            columns.append({"label": period, "subheaders": subheaders})
    else:
        for period in period_order:
            columns.append({"label": period, "subheaders": None})

    return columns, has_subheaders


def _period_entry_data(table_root: dict[str, Any], data_path):
    if not data_path:
        return None
    cur = table_root
    for key in data_path:
        if cur is None:
            return None
        cur = cur.get(key)
    return cur


def _build_period_row_cells(entry: Any, columns: list[dict]):
    cells = []
    for column in columns:
        if column["subheaders"]:
            period_values = entry.get(column["label"], {}) if isinstance(entry, dict) else {}
            cell_values = []
            for sub in column["subheaders"]:
                value = ""
                if isinstance(period_values, dict):
                    value = _format_display_value(period_values.get(sub["key"]))
                align = "right" if _is_numeric(period_values.get(sub["key"])) else "left"
                cell_values.append({"value": value, "align": align})
            cells.append({"type": "multi", "values": cell_values})
        else:
            value = ""
            if isinstance(entry, dict):
                raw = entry.get(column["label"])
                value = _format_display_value(raw)
            else:
                raw = None
            align = "right" if _is_numeric(raw) else "left"
            cells.append({"type": "single", "value": value, "align": align})
    return cells


def _build_structured_sections(records: list[dict]) -> list[dict]:
    sheets: "OrderedDict[str, OrderedDict[str, list]]" = OrderedDict()
    for record in records:
        labels = [token for token in record.get("label_path", []) if token]
        if not labels:
            continue
        sheet_name = labels[0]
        section_name = labels[1] if len(labels) > 1 else "세부 항목"
        entry_labels = labels[2:]

        sheet_entry = sheets.setdefault(sheet_name, OrderedDict())
        section_entry = sheet_entry.setdefault(section_name, [])
        section_entry.append(
            {
                "labels": entry_labels,
                "value": record.get("value"),
            }
        )

    structured = []
    for sheet_name, sections in sheets.items():
        structured_sections = []
        for section_name, entries in sections.items():
            structured_sections.append(_build_section_block(section_name, entries))
        structured.append({"sheet": sheet_name, "sections": structured_sections})
    return structured


def _build_section_block(section_name: str, entries: list[dict]) -> dict:
    has_multi_depth = any(len(entry["labels"]) >= 2 for entry in entries if entry.get("labels"))
    if not has_multi_depth:
        rows = []
        for idx, entry in enumerate(entries, start=1):
            labels = entry.get("labels") or []
            label = labels[0] if labels else f"{section_name} #{idx}"
            raw = entry.get("value")
            value_list = _as_list_items(raw)
            rows.append(
                {
                    "label": label,
                    "value": _format_display_value(raw),
                    "items": value_list,
                    "align": "right" if _is_numeric(raw) else "left",
                }
            )
        return {"mode": "list", "title": section_name, "rows": rows}

    return _build_table_block(section_name, entries)


def _build_table_block(section_name: str, entries: list[dict]) -> dict:
    row_order: list[str] = []
    col_order: list[str] = []
    table: "OrderedDict[str, OrderedDict[str, str]]" = OrderedDict()

    for entry in entries:
        labels = entry.get("labels") or []
        row_label = labels[0] if labels else section_name
        if len(labels) > 2:
            col_label = " / ".join(labels[1:])
        elif len(labels) == 2:
            col_label = labels[1]
        else:
            col_label = "값"

        row_label = str(row_label)
        col_label = str(col_label)

        if row_label not in table:
            table[row_label] = OrderedDict()
            row_order.append(row_label)
        if col_label not in col_order:
            col_order.append(col_label)
        table[row_label][col_label] = entry.get("value")

    rows = []
    for row_label in row_order:
        value_cells = []
        for col in col_order:
            raw = table[row_label].get(col, "")
            value_cells.append(
                {
                    "value": _format_display_value(raw),
                    "align": "right" if _is_numeric(raw) else "left",
                }
            )
        rows.append(
            {
                "row_label": row_label,
                "values": value_cells,
            }
        )

    return {
        "mode": "table",
        "title": section_name,
        "headers": col_order,
        "rows": rows,
    }


def create_fields_from_upload(submission: Submission, upload: Upload):
    """
    특정 업로드 파일을 기준으로 SubmissionField들을 생성/갱신
    - 기존 필드는 싹 지우고 새로 만든다
    """
    SubmissionField.objects.filter(submission=submission).delete()

    normalized_payload = {}
    normalized_records: list[dict] = []

    fallback_field_specs: list[dict] = []

    try:
        normalized_payload, normalized_records = parse_with_ntpd(upload.file.path)
    except NTPDParserUnavailable as exc:
        logger.warning("ntpd parser unavailable, fallback to simple extractor: %s", exc)
    except Exception as exc:
        logger.exception("ntpd parsing 실패 (%s): %s", upload.file.name, exc)
    finally:
        if not normalized_records:
            fallback_field_specs = extract_fields_from_upload(upload)
            normalized_payload = {}

    submission.normalized_payload = normalized_payload
    submission.normalized_records = normalized_records
    submission.save(update_fields=["normalized_payload", "normalized_records"])

    if normalized_records:
        overrides = submission.normalized_overrides or {}
        effective_records = _apply_overrides_to_records(normalized_records, overrides) if overrides else normalized_records
        _rebuild_submission_fields(submission, effective_records)
    else:
        SubmissionField.objects.filter(submission=submission).delete()
        for idx, spec in enumerate(fallback_field_specs, start=1):
            SubmissionField.objects.create(
                submission=submission,
                name=spec.get("name") or f"field_{idx}",
                label=spec.get("label") or f"Field {idx}",
                value=spec.get("value") or "",
                data_type=spec.get("data_type") or SubmissionField.DataType.TEXT,
                order=spec.get("order") or idx,
            )


# -------------------------------
# 뷰들
# -------------------------------

@login_required
def new_submission(request):
    """
    새 심의 신청 작성 (1단계)
    - 제목/설명 + 파일 업로드
    - 여기서는 무조건 DRAFT로 생성
    - 파일을 기반으로 SubmissionField 생성
    - 이후 edit_submission으로 리다이렉트하여 가변 필드/제출 처리
    """
    user = request.user
    if not getattr(user, "is_borrower", False):
        return HttpResponseForbidden("차주만 새로운 심의 신청을 만들 수 있습니다.")

    if request.method == "POST":
        sub_form = SubmissionForm(request.POST)
        upload_form = MultiUploadForm(request.POST, request.FILES)
        if sub_form.is_valid() and upload_form.is_valid():
            submission = sub_form.save(commit=False)
            submission.borrower = user
            submission.status = Submission.Status.DRAFT
            submission.save()

            # 상태변경 이벤트 기록
            SubmissionEvent.objects.create(
                submission=submission,
                actor=user,
                event_type=SubmissionEvent.EventType.STATUS_CHANGE,
                from_status=None,
                to_status=submission.status,
                field_name="general",
                message="초기 생성 및 저장",
            )

            # 파일 업로드
            uploads = []
            for i in range(1, 3):
                f = upload_form.cleaned_data.get(f"file_{i}")
                if f:
                    uploads.append(
                        Upload.objects.create(
                            submission=submission,
                            file=f,
                            original_name=f.name,
                        )
                    )

            # 파싱 기준이 될 파일 선택: 우선 file_1, 없으면 첫 업로드
            target_upload = uploads[0] if uploads else None

            if target_upload:
                create_fields_from_upload(submission, target_upload)

            return redirect("edit_submission", submission_id=submission.id)
    else:
        sub_form = SubmissionForm()
        upload_form = MultiUploadForm()

    return render(
        request,
        "submissions/new_submission.html",
        {
            "sub_form": sub_form,
            "upload_form": upload_form,
            "mode": "create",
        },
    )


@login_required
def edit_submission(request, submission_id: int):
    """
    차주가 기존 심의 신청을 수정하는 화면
    - 제목/설명 + 가변 필드들 + 파일목록/교체 + 타임라인
    - 임시저장: DRAFT
    - 검토 요청(제출): IN_REVIEW
    - 최종확정된 신청서는 수정 불가
    """
    user = request.user
    submission = get_object_or_404(Submission, id=submission_id)

    if submission.borrower != user:
        return HttpResponseForbidden("자신의 신청서만 수정할 수 있습니다.")

    fields_qs = submission.fields.all()

    if submission.status == Submission.Status.FINALIZED:
        if request.method == "POST":
            return HttpResponseForbidden("최종확정된 신청서는 더 이상 수정할 수 없습니다.")

    if request.method == "POST":
        sub_form = SubmissionForm(request.POST, instance=submission)
        upload_form = MultiUploadForm(request.POST, request.FILES)
        dynamic_form = SubmissionDynamicFieldsForm(request.POST, fields_qs=fields_qs)

        if sub_form.is_valid() and upload_form.is_valid() and dynamic_form.is_valid():
            old_status = submission.status
            submission = sub_form.save(commit=False)

            action = request.POST.get("action")
            if action == "submit":
                submission.status = Submission.Status.IN_REVIEW
            elif action == "draft":
                submission.status = Submission.Status.DRAFT

            submission.save()

            # 상태 변경 로그 남기기
            if old_status != submission.status:
                SubmissionEvent.objects.create(
                    submission=submission,
                    actor=user,
                    event_type=SubmissionEvent.EventType.STATUS_CHANGE,
                    from_status=old_status,
                    to_status=submission.status,
                    field_name="general",
                    message="차주가 신청서를 수정/제출했습니다.",
                )

            # 가변 필드 값 저장
            for field in fields_qs:
                form_field_name = SubmissionDynamicFieldsForm._field_name(field)
                value = dynamic_form.cleaned_data.get(form_field_name)
                field.value = "" if value is None else str(value)
                field.save()

            # 새로 업로드된 파일 추가
            for i in range(1, 3):
                f = upload_form.cleaned_data.get(f"file_{i}")
                if f:
                    Upload.objects.create(
                        submission=submission,
                        file=f,
                        original_name=f.name,
                    )

            return redirect("borrower_dashboard")
    else:
        sub_form = SubmissionForm(instance=submission)
        upload_form = MultiUploadForm()
        dynamic_form = SubmissionDynamicFieldsForm(fields_qs=fields_qs)

    events = submission.events.select_related("actor").all()

    return render(
        request,
        "submissions/edit_submission.html",
        {
            "submission": submission,
            "sub_form": sub_form,
            "upload_form": upload_form,
            "dynamic_form": dynamic_form,
            "mode": "edit",
            "events": events,
        },
    )


@login_required
def replace_upload_inline(request, upload_id: int):
    """
    차주가 기존 파일을 교체
    """
    upload = get_object_or_404(Upload, id=upload_id)
    user = request.user

    if upload.submission.borrower != user:
        return HttpResponseForbidden("자신의 파일만 교체할 수 있습니다.")

    if upload.submission.status == Submission.Status.FINALIZED:
        return HttpResponseForbidden("최종확정된 신청서의 파일은 교체할 수 없습니다.")

    if request.method == "POST" and request.FILES.get("file"):
        new_file = request.FILES["file"]
        upload.file = new_file
        upload.original_name = new_file.name
        upload.save()

        SubmissionEvent.objects.create(
            submission=upload.submission,
            actor=user,
            event_type=SubmissionEvent.EventType.COMMENT,
            field_name="general",
            message=f"파일 '{upload.original_name}' 을 교체했습니다.",
        )

    return HttpResponseRedirect(reverse("borrower_dashboard"))


@login_required
def submission_review(request, submission_id: int):
    """
    대주가 신청서 상세를 보고
    - 코멘트/수정요청 남기기 (드롭다운 없이 코멘트만 입력)
    - 상태를 최종확정(FINALIZED)으로 변경
    - '수정요청' 버튼을 누르면 상태를 REVISION_REQUIRED로 변경
    """
    user = request.user
    if not _is_lender_like(user):
        return HttpResponseForbidden("대주만 접근할 수 있습니다.")

    submission = get_object_or_404(Submission, id=submission_id)

    if request.method == "POST":
        action = request.POST.get("action")

        # 1) 코멘트/수정요청 추가
        if action in ("comment", "request_revision"):
            form = SubmissionEventForm(request.POST)
            if form.is_valid():
                ev = form.save(commit=False)
                ev.submission = submission
                ev.actor = user
                ev.field_name = "general"

                if action == "request_revision":
                    # 수정요청이면 상태도 같이 변경
                    old_status = submission.status
                    submission.status = Submission.Status.REVISION_REQUIRED
                    submission.save()

                    ev.event_type = SubmissionEvent.EventType.REQUEST
                    ev.from_status = old_status
                    ev.to_status = submission.status
                else:
                    ev.event_type = SubmissionEvent.EventType.COMMENT

                ev.save()
                return redirect("submission_review", submission_id=submission.id)

        # 2) 최종확정
        elif action == "finalize":
            old_status = submission.status
            submission.status = Submission.Status.FINALIZED
            submission.save()

            SubmissionEvent.objects.create(
                submission=submission,
                actor=user,
                event_type=SubmissionEvent.EventType.STATUS_CHANGE,
                from_status=old_status,
                to_status=submission.status,
                field_name="general",
                message="대주가 최종확정했습니다.",
            )
            return redirect("submission_review", submission_id=submission.id)

    events = submission.events.select_related("actor").all()
    fields_qs = submission.fields.all()

    return render(
        request,
        "submissions/submission_review.html",
        {
            "submission": submission,
            "events": events,
            "event_form": SubmissionEventForm(),
            "fields": fields_qs,
        },
    )


@login_required
def submission_data_view(request, submission_id: int):
    """
    차주/대주 모두가 사용할 수 있는 ntpd 추출 데이터 미리보기
    - 시트 단위 → 섹션 단위 → 표 또는 리스트 형태로 렌더링
    """
    submission = get_object_or_404(Submission, id=submission_id)
    user = request.user

    if submission.borrower != user and not _is_lender_like(user):
        return HttpResponseForbidden("접근 권한이 없습니다.")

    payload = _get_effective_payload(submission)
    records = _get_effective_records(submission)

    spec_views = _build_spec_views_from_payload(payload)
    sheet_map: "OrderedDict[str, list]" = OrderedDict(
        (view["sheet"], list(view["sections"])) for view in spec_views
    )

    generic_views = _build_structured_sections(records) if records else []
    for view in generic_views:
        if view["sheet"] in sheet_map:
            continue
        sheet_map.setdefault(view["sheet"], []).extend(view["sections"])

    sheet_views = [
        {"sheet": sheet_name, "sections": sections}
        for sheet_name, sections in sheet_map.items()
        if sections
    ]

    back_url = None
    back_label = None
    if submission.borrower == user:
        back_url = reverse("edit_submission", args=[submission.id])
        back_label = "신청서 수정으로 돌아가기"
    elif _is_lender_like(user):
        back_url = reverse("submission_review", args=[submission.id])
        back_label = "대주 검토 화면으로 돌아가기"

    return render(
        request,
        "submissions/submission_data_view.html",
        {
            "submission": submission,
            "records_count": len(records),
            "sheet_views": sheet_views,
            "has_records": bool(sheet_views),
            "back_url": back_url,
            "back_label": back_label,
        },
    )


@login_required
def edit_submission_data(request, submission_id: int):
    submission = get_object_or_404(Submission, id=submission_id)
    user = request.user

    if submission.borrower != user:
        return HttpResponseForbidden("차주만 데이터를 수정할 수 있습니다.")
    if submission.status == Submission.Status.FINALIZED:
        return HttpResponseForbidden("최종확정된 신청서는 수정할 수 없습니다.")
    if not submission.normalized_records:
        return HttpResponseForbidden("수정 가능한 데이터가 없습니다.")

    base_records = submission.normalized_records or []
    overrides = submission.normalized_overrides or {}
    base_map = {}
    initial = []

    for rec in base_records:
        key = _key_path_to_str(rec.get("key_path"))
        label = " / ".join([str(x) for x in rec.get("label_path") or []])
        base_value = str(rec.get("value") or "")
        current_value = str(overrides.get(key, base_value))
        base_map[key] = base_value
        initial.append(
            {
                "key": key,
                "label": label or "(라벨 없음)",
                "value": current_value,
            }
        )

    SubmissionDataFormSet = formset_factory(SubmissionDataForm, extra=0)
    prefix = "records"

    if request.method == "POST":
        formset = SubmissionDataFormSet(request.POST, prefix=prefix)
        if formset.is_valid():
            new_overrides = {}
            for form in formset:
                key = form.cleaned_data.get("key")
                value = form.cleaned_data.get("value") or ""
                base_value = base_map.get(key, "")
                if value != base_value:
                    new_overrides[key] = value
            submission.normalized_overrides = new_overrides
            submission.save(update_fields=["normalized_overrides"])

            effective_records = (
                _apply_overrides_to_records(base_records, new_overrides)
                if new_overrides
                else base_records
            )
            if base_records:
                _rebuild_submission_fields(submission, effective_records)

            messages.success(request, "데이터를 저장했습니다.")
            return redirect("edit_submission_data", submission_id=submission.id)
    else:
        formset = SubmissionDataFormSet(initial=initial, prefix=prefix)

    return render(
        request,
        "submissions/submission_data_edit.html",
        {
            "submission": submission,
            "formset": formset,
        },
    )
class SubmissionDataForm(forms.Form):
    key = forms.CharField(widget=forms.HiddenInput)
    label = forms.CharField(widget=forms.HiddenInput)
    value = forms.CharField(
        widget=forms.Textarea(attrs={"rows": 2, "style": "width:100%;"}),
        required=False,
    )
